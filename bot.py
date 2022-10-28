import requests
from bs4 import BeautifulSoup
from openpyxl import *
from openpyxl.styles import Font, Alignment
import googlemaps

url = "https://moovitapp.com/index/pt-br/transporte_p%C3%BAblico-line-0903-Maceio-4466-2221160-91572044-3"
page = requests.get(url)
soup = BeautifulSoup(page.content, 'html.parser')


def get_stops_addresses():
    stops_container = soup.find_all('li', class_='stop-container')
    stops_wrappers = [stop.find('div', class_='stop-wrapper') for stop in stops_container]
    stops_addresses = []
    for stop in stops_wrappers:
        stop_address = stop.find('h3').text
        stops_addresses.append(stop_address)
    return stops_addresses


def get_line_number():
    title = soup.find('div', class_='line-image-container')
    line_number = title.find('h1', class_="text").text

    return line_number


def get_title(line_number):
    line_title = soup.find('div', class_='line-title').find('h2', class_="title").text

    return line_number + " - " + line_title


def get_coordinates(stops):
    gmaps = googlemaps.Client(key='YOUR_API_KEY')
    coordinates = []
    for stop in stops:
        geocode_result = gmaps.geocode(stop)
        if geocode_result:
            coordinates.append(geocode_result[0]['geometry']['location'])
        else:
            coordinates.append(None)
    return coordinates


def save_into_sheet():
    print("Getting stops addresses...")
    sheets = Workbook()
    sheet = sheets.active

    line_number = get_line_number()
    stops = get_stops_addresses()
    coordinates = get_coordinates(stops)

    # increase the size of columns
    sheet.column_dimensions['A'].width = 150
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30

    sheet.cell(row=1, column=1).value = get_title(line_number)
    sheet.cell(row=1, column=1).font = Font(bold=True, size=20)
    sheet.cell(row=1, column=2).value = "Latitude"
    sheet.cell(row=1, column=2).font = Font(bold=True, size=20)
    sheet.cell(row=1, column=3).value = "Longitude"
    sheet.cell(row=1, column=3).font = Font(bold=True, size=20)
    
    for i in range(len(stops)):
        sheet.cell(row=i + 2, column=1).value = stops[i]
        if coordinates[i]:
            sheet.cell(row=i + 2, column=2).value = coordinates[i]['lat']
            sheet.cell(row=i + 2, column=3).value = coordinates[i]['lng']

            sheet.cell(row=i + 2, column=2).alignment = Alignment(horizontal='left') 
            sheet.cell(row=i + 2, column=3).alignment = Alignment(horizontal='left')

    sheets.save(f'linhas/{line_number}.xlsx')
    print(f"Saved {line_number}.xlsx")


save_into_sheet()